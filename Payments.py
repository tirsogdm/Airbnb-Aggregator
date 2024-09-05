from PyQt5.QtWidgets import QTableView, QHeaderView, QWidget, QVBoxLayout, QFileDialog, QMessageBox
from PyQt5.QtCore import Qt, QAbstractTableModel, QSortFilterProxyModel, QDateTime, QDate, QTime
from Controller import Controller
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle
import os

from Representations import Amount


class Payments(QWidget):
    def __init__(self, data):
        super().__init__()
        self._data = data

        layout = QVBoxLayout()
        self.controller = Controller()
        self.payments_table = PaymentsTable(data)

        layout.addWidget(self.controller)
        layout.addWidget(self.payments_table)

        self.setLayout(layout)

        # Connections
        self.controller.search_filter.textChanged.connect(
            self.payments_table.proxy_filter_model.setFilterFixedString)
        self.controller.signals.update.connect(self.handle_date_selection)
        self.controller.export_btn.clicked.connect(self.generate_csv)

        self.controller.trigger_initial_signals()

    def handle_date_selection(self, year, month):
        min_date, max_date = self.get_min_max_dates(year, month)
        self.payments_table.proxy_filter_model.set_date_range(
            min_date, max_date)
        self.payments_table.clearSelection()

    def get_min_max_dates(self, year, month):
        min_date = QDate(year, month, 1)
        min_date_time = QDateTime(min_date, QTime(0, 0, 0))

        days_in_month = min_date.daysInMonth()

        max_date_time = QDateTime(
            QDate(year, month, days_in_month), QTime(23, 59, 59))

        return min_date_time, max_date_time

    def get_visible_rows(self):
        row_count = self.payments_table.proxy_filter_model.rowCount()

        visible_rows = list()
        for row in range(row_count):
            proxy_index = self.payments_table.proxy_filter_model.index(
                row, 0)
            source_index = self.payments_table.proxy_filter_model.mapToSource(
                proxy_index)
            visible_rows.append(source_index.row())

        visible_rows.sort()  # set to ascending order by default
        return visible_rows

    def generate_csv(self):
        date = self.controller.get_date()
        file_name = f"Pagos Madrid Rentals - {date.month()}.{date.year()}.xlsx"

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        directory = QFileDialog.getExistingDirectory(
            self, "Select Directory", options=options)

        if directory:
            file_path = os.path.join(directory, file_name)
            self.create_workbook(file_path)

            QMessageBox.information(
                self, "Success", f"Excel file created successfully at:\n{file_path}")

    def create_workbook(self, path):
        visible_rows = self.get_visible_rows()

        wb = Workbook()

        # Create worksheets
        ws_ch7 = wb.active
        ws_ch7.title = "CH7"

        ws_v41 = wb.create_sheet("V41")

        # Currency format for Euro
        euro_format = NamedStyle(name="euro_format")
        euro_format.number_format = '#,##0.00 €'

        # Add headers
        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        headers = ["", "Platform", "Date", "Amount Transfered",
                   "Amount", "Guest", "Apartment", "Reference"]

        ws_ch7.append(headers)
        ws_v41.append(headers)

        for row in visible_rows:
            payment = self._data[row]
            for i, reservation in enumerate(payment.reservations):
                row = ["", "Airbnb", "", "", reservation.amount,
                       reservation.guest_name, reservation.apartment, reservation.id]

                if reservation.building == "CH7":
                    target_sheet = ws_ch7
                elif reservation.building == "V41":
                    target_sheet = ws_v41
                else:
                    # Error checking - Think of how to expose this to user.
                    print("NOT IN ANY BUILDING!!")
                    continue

                if i == 0:
                    row[2] = payment.date.toString("dd/MM/yyyy")
                    row[3] = payment.amount

                    target_sheet.append(row)
                    for cell in target_sheet[target_sheet.max_row]:
                        cell.fill = header_fill

                    target_sheet[target_sheet.max_row][3].style = euro_format

                else:
                    target_sheet.append(row)

                target_sheet[target_sheet.max_row][4].style = euro_format

        # Size columns
        for ws in [ws_ch7, ws_v41]:
            for column in ws.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                adjusted_width = max_length + 2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(path)


class PaymentsTable(QTableView):
    def __init__(self, data):
        super().__init__()

        # Table options
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.setSelectionMode(QTableView.SingleSelection)
        self.setSelectionBehavior(QTableView.SelectRows)
        self.setSortingEnabled(True)
        self.verticalHeader().setVisible(False)

        self.model = PaymentsModel(data)

        self.proxy_filter_model = PaymentsSortFilterProxyModel()
        self.proxy_filter_model.setSourceModel(self.model)
        self.proxy_filter_model.setFilterKeyColumn(-1)

        self.setModel(self.proxy_filter_model)

        """
        self.proxy_filter_model.set_building_filter("CH7")
        self.proxy_filter_model.set_date_range(
            QDateTime.fromString('Tue, 23 Jul 2024 14:00:00',
                                 'ddd, dd MMM yyyy HH:mm:ss'),
            QDateTime.fromString('Sun, 28 Jul 2024 14:00:00', 'ddd, dd MMM yyyy HH:mm:ss'))
        """


class PaymentsModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data

    def data(self, index, role):
        value = self._data[index.row()]
        idx_col = index.column()

        if role == Qt.DisplayRole:
            if idx_col == 0:
                return value.datetime.toString()
            elif idx_col == 1:
                return value.amount.toString()
            elif idx_col == 2:
                return value.building
            elif idx_col == 3:
                return value.num_reservations

        if role == Qt.EditRole or role == Qt.UserRole:
            if idx_col == 0:
                return value.datetime
            elif idx_col == 1:
                return value.amount
            elif idx_col == 2:
                return value.building
            elif idx_col == 3:
                return value.num_reservations

        if role == Qt.TextAlignmentRole:
            return Qt.AlignCenter

    def rowCount(self, index):
        return len(self._data)

    def columnCount(self, index):
        return 4

    def headerData(self, section, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            # Set custom horizontal headers
            headers = ["Date", "Amount", "Building",
                       "Reservations"]
            return headers[section]

        return super().headerData(section, orientation, role)


class PaymentsSortFilterProxyModel(QSortFilterProxyModel):
    def __init__(self):
        super(PaymentsSortFilterProxyModel, self).__init__()
        self.min_date = None
        self.max_date = None

    def set_date_range(self, min_date, max_date):
        self.min_date = min_date
        self.max_date = max_date
        self.invalidateFilter()

    def lessThan(self, left, right):
        left_data = self.sourceModel().data(left, Qt.EditRole)
        right_data = self.sourceModel().data(right, Qt.EditRole)

        if isinstance(left_data, QDateTime) and isinstance(right_data, QDateTime):
            return left_data < right_data

        if isinstance(left_data, Amount) and isinstance(right_data, Amount):
            return left_data < right_data

        return super(PaymentsSortFilterProxyModel, self).lessThan(left, right)

    def filterAcceptsRow(self, source_row, source_parent):
        model = self.sourceModel()

        if self.min_date is not None or self.max_date is not None:
            date_index = model.index(source_row, 0)
            date_value = model.data(date_index, Qt.EditRole)

            if (self.min_date is not None and date_value < self.min_date) or \
                    (self.max_date is not None and date_value > self.max_date):
                return False

        return super(PaymentsSortFilterProxyModel, self).filterAcceptsRow(source_row, source_parent)
